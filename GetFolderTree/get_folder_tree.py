import os
import sys

from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLineEdit, QPushButton, QLabel, QHBoxLayout, QFileDialog, QMessageBox, QCheckBox
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException


class FolderTreeApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('文件目录树生成工具')
        self.setGeometry(100, 100, 400, 200)
        self.setAcceptDrops(True)

        layout = QVBoxLayout()

        # 处理目录
        self.source_path_label = QLabel('处理目录:')
        layout.addWidget(self.source_path_label)
        source_path_layout = QHBoxLayout()
        self.source_path_input = QLineEdit(self)
        self.source_path_input.setAcceptDrops(True)
        self.source_path_input.dragEnterEvent = self.dragEnterEvent
        self.source_path_input.dropEvent = self.dropEvent
        self.source_path_input.textChanged.connect(self.validate_source_path)
        source_path_layout.addWidget(self.source_path_input)
        self.source_path_browse_button = QPushButton('浏览', self)
        self.source_path_browse_button.clicked.connect(self.browse_source_path)
        source_path_layout.addWidget(self.source_path_browse_button)
        layout.addLayout(source_path_layout)

        self.scan_button = QPushButton('扫描目录', self)
        self.scan_button.clicked.connect(self.scan_directory)
        layout.addWidget(self.scan_button)

        # 输出格式
        self.format_label = QLabel('输出格式:')
        layout.addWidget(self.format_label)
        self.format_input = QLineEdit(self)
        self.format_input.textChanged.connect(self.validate_format_string)
        layout.addWidget(self.format_input)
        self.format_input.setText('$name$')  # 设置默认值

        # 输出层级
        self.level_label = QLabel('输出层级:')
        layout.addWidget(self.level_label)
        self.level_input = QLineEdit(self)
        self.level_input.textChanged.connect(self.validate_level)
        layout.addWidget(self.level_input)
        self.level_input.setText('100')  # 设置默认值

        # 保存目录
        self.save_path_label = QLabel('保存目录:')
        layout.addWidget(self.save_path_label)
        save_path_layout = QHBoxLayout()
        self.save_path_input = QLineEdit(self)
        self.save_path_input.textChanged.connect(self.validate_save_path)
        save_path_layout.addWidget(self.save_path_input)
        self.save_path_browse_button = QPushButton('浏览', self)
        self.save_path_browse_button.clicked.connect(self.browse_save_path)
        save_path_layout.addWidget(self.save_path_browse_button)
        layout.addLayout(save_path_layout)

        # 添加复选框
        self.output_files_checkbox = QCheckBox('输出文件', self)
        layout.addWidget(self.output_files_checkbox)

        self.save_markdown_button = QPushButton('保存Markdown', self)
        self.save_markdown_button.clicked.connect(self.save_markdown)
        layout.addWidget(self.save_markdown_button)

        self.save_xlsx_button = QPushButton('保存XLSX', self)
        self.save_xlsx_button.clicked.connect(self.save_xlsx)
        layout.addWidget(self.save_xlsx_button)

        self.setLayout(layout)

    def validate_source_path(self):
        source_path = self.source_path_input.text()
        if source_path and not os.path.isdir(source_path):
            QMessageBox.critical(self, '错误', '处理目录无效')

    def validate_save_path(self):
        save_path = self.save_path_input.text()
        if save_path and not os.path.isdir(save_path):
            QMessageBox.critical(self, '错误', '保存目录无效')

    def validate_format_string(self):
        format_string = self.format_input.text()
        if format_string and '$name$' not in format_string:
            QMessageBox.critical(self, '错误', '格式字符串必须包含 $name$')

    def validate_level(self):
        level_string = self.level_input.text()
        try:
            if level_string:
                int(level_string)
        except ValueError:
            QMessageBox.critical(self, '错误', '输出层级必须是整数')

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            self.source_path_input.setText(urls[0].toLocalFile())
            QMessageBox.information(self, '成功', f'添加目录：{urls[0].toLocalFile()}')

    def browse_source_path(self):
        directory = QFileDialog.getExistingDirectory(self, "选择处理目录")
        if directory:
            self.source_path_input.setText(directory)

    def browse_save_path(self):
        directory = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if directory:
            self.save_path_input.setText(directory)

    def scan_directory(self):
        source_path = self.source_path_input.text()
        if os.path.isdir(source_path):
            print(f"Scanning directory: {source_path}")
            try:
                self.folder_tree = get_file_folder_tree(source_path)
                QMessageBox.information(self, '成功', '扫描完成')
                print("Scan completed")
            except Exception as e:
                QMessageBox.critical(self, '错误', f'扫描目录时出错：{e}')
                print(f"Error while scanning directory: {e}")
        else:
            QMessageBox.critical(self, '错误', '无效目录')
            print("Invalid directory")

    def save_markdown(self):
        save_path = self.save_path_input.text()
        format_string = self.format_input.text()
        if save_path and format_string:
            print(f"Saving Markdown to: {save_path} with format: {format_string}")
            try:
                print_markdown_list(self.output_files_checkbox.isChecked(), format_string, int(self.level_input.text()), self.folder_tree,
                                    (save_path + '/FileFolderTree.md') if self.output_files_checkbox.isChecked() else (save_path + '/FolderTree.md'))
                QMessageBox.information(self, '成功', '保存完成')
                print("Save completed")
            except Exception as e:
                QMessageBox.critical(self, '错误', f'保存 Markdown 时出错：{e}')
                print(f"Error while saving Markdown file: {e}")
        else:
            QMessageBox.critical(self, '错误', '无效保存目录或格式字符串')
            print("Invalid save path or format string")

    def save_xlsx(self):
        save_path = self.save_path_input.text()
        format_string = self.format_input.text()
        if save_path and format_string:
            print(f"Saving XLSX to: {save_path} with format: {format_string}")
            try:
                print_xlsx(self.output_files_checkbox.isChecked(), format_string, int(self.level_input.text()), self.folder_tree,
                           (save_path + '/FileFolderTree.xlsx') if self.output_files_checkbox.isChecked() else (save_path + '/FolderTree.xlsx'))
                QMessageBox.information(self, '成功', '保存完成')
                print("Save completed")
            except Exception as e:
                QMessageBox.critical(self, '错误', f'保存 XLSX 时出错：{e}')
                print(f"Error while saving XLSX file: {e}")
        else:
            QMessageBox.critical(self, '错误', '无效保存目录或格式字符串')
            print("Invalid save path or format string")


def get_file_folder_tree(source_path):
    # 获取指定目录的结构，包括文件夹和文件，以树的形式存储
    folder_tree = {}
    try:
        # 使用 os.walk 遍历目录树
        for root, dirs, files in os.walk(source_path):
            # 获取相对于 source_path 的相对路径
            relative_path = os.path.relpath(root, source_path)
            current_level = folder_tree

            # 如果相对路径不是当前目录
            if relative_path != '.':
                # 根据路径的各个部分逐级深入字典
                for part in relative_path.split(os.sep):
                    current_level = current_level.setdefault(part, {})

            # 将子目录添加到当前级别的字典中
            for dir_name in dirs:
                current_level[dir_name] = {}
            # 将文件添加到当前级别的字典中，值为 None
            for file_name in files:
                print(f"File: {file_name}")
                current_level[file_name] = None
    except Exception as e:
        print(f"Error while reading directory structure: {e}")
    return folder_tree


def is_file(path):
    # 分离文件名和扩展名
    _, ext = os.path.splitext(path)
    # 如果有扩展名，则认为是文件
    return bool(ext)


def print_markdown_list(print_files, format_string, level, folder_tree, result_path):
    # 遍历树，根据层级输出 markdown 列表
    # format_string："$name$" 代表文件名
    def traverse_and_generate_markdown(tree, current_level=0):
        if current_level > level - 1:
            return []
        markdown_lines = []
        for name, subtree in tree.items():
            if print_files is True or is_file(name) is False:
                # 根据层级生成缩进
                indent = '  ' * current_level
                # 格式化文件名
                formatted_name = format_string.replace('$name$', name)
                # 添加到 markdown 列表
                markdown_lines.append(f'{indent}{formatted_name}')
            # 如果是子目录，递归处理
            if isinstance(subtree, dict):
                markdown_lines.extend(traverse_and_generate_markdown(subtree, current_level + 1))
        return markdown_lines

    try:
        # 生成 markdown 列表
        markdown_list = traverse_and_generate_markdown(folder_tree)
        # 将结果写入文件
        with open(result_path, 'w', encoding='utf-8') as file:
            file.write('\n'.join(markdown_list))
    except IOError as e:
        print(f"Error while writing to markdown file: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")


def print_xlsx(print_files, format_string, level, folder_tree, result_path):
    # 遍历树，根据层级输出 XLSX 文件，合并单元格表示层级关系
    # format_string："$name$" 代表文件名
    def traverse_and_generate_xlsx(tree, ws, row=1, col=1):
        for name, subtree in tree.items():
            if print_files is True or os.path.is_file(name) is False:
                # 格式化文件名
                formatted_name = format_string.replace('$name$', name)
                # 在工作表中写入文件名
                cell = ws.cell(row=row, column=col, value=formatted_name)
                # 设置单元格水平居中、垂直居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                row -= 1
            # 如果是子目录，递归处理
            if subtree is not None and len(subtree) > 0:
                next_row = traverse_and_generate_xlsx(subtree, ws, row, col + 1)
                # 合并单元格表示层级关系
                ws.merge_cells(start_row=row, start_column=col, end_row=next_row - 1, end_column=col)
                row = next_row
            else:
                row += 1
        return row

    try:
        # 创建一个新的工作簿
        wb = Workbook()
        # 获取活动工作表
        ws = wb.active
        # 生成 XLSX 文件内容
        traverse_and_generate_xlsx(folder_tree, ws)
        # 调整列宽以适应内容
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        # 保存工作簿到文件
        wb.save(result_path)
    except InvalidFileException as e:
        print(f"Invalid file error: {e}")
    except PermissionError as e:
        print(f"Permission error: {e}")
    except Exception as e:
        print(f"Error while generating or saving XLSX file: {e}")


#         TODO：不输出文件时会报错Min value is 1，还没有实现输出指定层级的功能


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('icon.ico'))
    ex = FolderTreeApp()
    ex.show()
    sys.exit(app.exec_())

# pyinstaller --onefile --windowed --icon=icon.ico get_folder_tree.py
