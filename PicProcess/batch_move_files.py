import os
import shutil
from fractions import Fraction

import openpyxl
import rawpy
from PIL import Image
from PIL.ExifTags import TAGS
from PIL.TiffImagePlugin import IFDRational


def get_image_metadata(file_path):
    metadata = {
        '拍摄日期': '',
        '修改日期': '',
        '分级': '',
        '文件大小': 0,
        '宽度': '',
        '高度': '',
        '相机制造商': '',
        '相机型号': ''
    }

    try:
        metadata['文件大小'] = os.path.getsize(file_path) / (1024 * 1024)  # Set file size in MB
        extension = os.path.splitext(file_path)[1].lower()
        if extension in ['.arw', '.nef', '.rw2', '.dng']:
            with rawpy.imread(file_path) as raw:
                metadata['宽度'], metadata['高度'] = raw.sizes.width, raw.sizes.height
                metadata['相机制造商'] = raw.camera_make
                metadata['相机型号'] = raw.camera_model
                # Add more metadata extraction as needed
        else:
            image = Image.open(file_path)
            exif_data = image._getexif()
            if exif_data:
                for tag, value in exif_data.items():
                    tag_name = TAGS.get(tag, tag)
                    if isinstance(value, bytes):
                        value = value.decode(errors='ignore')
                    elif isinstance(value, (int, float, Fraction, IFDRational)):
                        value = str(value)
                    if tag_name == 'DateTimeOriginal':
                        metadata['拍摄日期'] = value
                    elif tag_name == 'DateTime':
                        metadata['修改日期'] = value
                    elif tag_name == 'Rating':
                        metadata['分级'] = value
                    elif tag_name == 'Make':
                        metadata['相机制造商'] = value
                    elif tag_name == 'Model':
                        metadata['相机型号'] = value
            metadata['宽度'], metadata['高度'] = image.size
    except Exception as e:
        print(f"Error reading metadata from {file_path}: {e}")

    return metadata


def get_file_names(target_dir, process_sub_folder=False, xlsx_file='file_names.xlsx'):
    """List all files in the target directory and save their names and extensions to an Excel file."""
    file_names = []

    if process_sub_folder:
        for root, _, files in os.walk(target_dir):
            for file in files:
                file_names.append(os.path.join(root, file))
    else:
        all_items = os.listdir(target_dir)
        file_names = [os.path.join(target_dir, item) for item in all_items if os.path.isfile(os.path.join(target_dir, item))]

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['文件名', '拓展名', '拍摄日期', '修改日期', '分级', '文件大小', '宽度', '高度', '相机制造商', '相机型号', '绝对路径']
    ws.append(headers)

    for index, file_path in enumerate(file_names, start=2):
        base_name, extension = os.path.splitext(os.path.basename(file_path))
        row = [base_name, extension]

        if extension.lower() in ['.jpg', '.jpeg', '.png', '.heic', '.arw', '.nef', '.rw2', '.dng']:
            metadata = get_image_metadata(file_path)
            row.extend([
                metadata['拍摄日期'],
                metadata['修改日期'],
                metadata['分级'],
                metadata['文件大小'],
                metadata['宽度'],
                metadata['高度'],
                metadata['相机制造商'],
                metadata['相机型号']
            ])
        else:
            row.extend([''] * 8)

        row.append(file_path)  # Add the absolute file path
        ws.append(row)

    wb.save(os.path.join(target_dir, xlsx_file))


def move_files(file_name_table, target_dir):
    """Move files listed in the Excel file to the target directory."""
    wb = openpyxl.load_workbook(file_name_table)
    ws = wb.active

    if not os.path.exists(target_dir):
        os.makedirs(target_dir)

    for row in ws.iter_rows(min_row=2, values_only=True):
        file_name, extension, _, _, _, _, _, _, _, _, absolute_path = row
        if not extension.startswith('.'):
            extension = '.' + extension
        full_file_name = f"{file_name}{extension}"

        target_path = os.path.join(target_dir, full_file_name)

        if os.path.exists(absolute_path):
            shutil.move(absolute_path, target_path)
